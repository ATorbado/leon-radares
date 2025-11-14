import csv, json, io, urllib.request, datetime
from pathlib import Path
from openpyxl import load_workbook

PROVINCIA = "LEON"
DGT_XLSX_URL = "https://www.dgt.es/export/sites/web-DGT/.galleries/downloads/conoce-el-estado-del-trafico/informacion-e-incidencias-de-trafico/JO_INFORME_CINEMOMETROS_WEB_DGT.XLSX"

def leer_xlsx(url):
    data = urllib.request.urlopen(url, timeout=60).read()
    wb = load_workbook(io.BytesIO(data), data_only=True)
    sh = wb.active

    headers = [
        (str(c.value).strip() if c.value is not None else "")
        for c in next(sh.iter_rows(min_row=1, max_row=1))
    ]

    rows = []
    for r in sh.iter_rows(min_row=2):
        row = {
            headers[i]: ("" if r[i].value is None else str(r[i].value))
            for i in range(len(headers))
        }
        rows.append(row)
    return rows

def filtrar(rows):
    out = []
    for r in rows:
        provincia = str(r.get("PROVINCIA", r.get("Provincia", ""))).strip().upper()
        if provincia != PROVINCIA:
            continue

        tipo = str(r.get("TIPO", r.get("Tipo", ""))).strip().upper()
        if "FIJO" not in tipo:
            continue

        out.append({
            "provincia": provincia.title(),
            "carretera": r.get("CARRETERA", r.get("Vía", "")),
            "tipo": tipo.title(),
            "pk": r.get("PK", ""),
            "sentido": r.get("SENTIDO", r.get("Sentido", "")),
            "last_update": datetime.datetime.utcnow().isoformat() + "Z"
        })

    return out

def main():
    rows = leer_xlsx(DGT_XLSX_URL)
    radares = filtrar(rows)
    Path("radars").mkdir(exist_ok=True)
    Path("radars/radares_fijos_sin_coord.json").write_text(
        json.dumps(radares, ensure_ascii=False, indent=2)
    )

if __name__ == "__main__":
    main()
